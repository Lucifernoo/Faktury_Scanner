from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from core.config_manager import DEFAULT_SETTINGS, ConfigManager

_DEFAULT_NAME = "Реєстр_рахунків.xlsx"


def _resolve_filepath(filepath: str) -> str:
    if os.path.isabs(filepath):
        return filepath
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    return os.path.join(root, filepath)


def get_registry_path(filename: str = _DEFAULT_NAME) -> str:
    """Абсолютний шлях до реєстру рахунків (як у ``ExcelExporter``)."""
    return os.path.abspath(_resolve_filepath(filename))


class ExcelExporter:
    """Дописує рядки в Excel; заголовки та порядок колонок — з ``export_columns`` (або settings.json)."""

    def __init__(
        self,
        config_manager: ConfigManager | None = None,
        file_path: str | None = None,
    ) -> None:
        self._config_mgr = config_manager or ConfigManager()
        if file_path:
            self.file_path = os.path.abspath(_resolve_filepath(file_path))
        else:
            self.file_path = get_registry_path()

    @staticmethod
    def _normalize_header(header: str) -> str:
        return re.sub(r"\s+", " ", header.strip().lower())

    @classmethod
    def classify_export_column(cls, header: str) -> str:
        """
        Зіставляє заголовок колонки зі слотом: date / vendor / edrpou / iban / amount / unknown.

        Порядок важливий: «ЄДРПОУ Контрагента» має ``єдрпоу`` і ``контрагент`` — спочатку перевіряємо ЄДРПОУ,
        і лише потім «чистий» контрагент без ЄДРПОУ в назві колонки.
        """
        h = cls._normalize_header(header)
        if "дата" in h and "скан" in h:
            return "date"
        # ПРАВИЛО 1: будь-яка колонка з «єдрпоу» — лише код ЄДРПОУ
        if "єдрпоу" in h or "є дрпоу" in h:
            return "edrpou"
        # ПРАВИЛО 2: контрагент у назві, але не колонка про ЄДРПОУ
        if "контрагент" in h and "єдрпоу" not in h:
            return "vendor"
        if "iban" in h or "ібан" in h:
            return "iban"
        if "сума" in h or "сплат" in h or h.endswith("грн") or "разом" in h:
            return "amount"
        return "unknown"

    def _build_slot_values(self, data: dict[str, Any]) -> dict[str, Any]:
        scan_dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        edrpou_cell = ", ".join(str(x) for x in (data.get("edrpou") or []))
        iban_cell = ", ".join(str(x) for x in (data.get("iban") or []))
        total = data.get("total_amount")
        if total is None or total == "":
            total_cell: str | float = ""
        else:
            try:
                total_cell = float(total)
            except (TypeError, ValueError):
                total_cell = str(total)
        vendor = str(data.get("vendor_name") or data.get("counterparty_name") or "").strip()
        return {
            "date": scan_dt,
            "vendor": vendor,
            "edrpou": edrpou_cell,
            "iban": iban_cell,
            "amount": total_cell,
        }

    def build_row_for_columns(
        self,
        data: dict[str, Any],
        export_columns: list[str] | None,
    ) -> tuple[list[str], list[Any]]:
        cols = list(export_columns) if export_columns else list(self._config_mgr.load().get("export_columns") or DEFAULT_SETTINGS["export_columns"])
        slots = self._build_slot_values(data)
        row: list[Any] = []
        for h in cols:
            slot = self.classify_export_column(h)
            if slot == "unknown":
                row.append("")
            else:
                row.append(slots.get(slot, ""))
        return cols, row

    def append_to_excel(
        self,
        data: dict[str, Any],
        export_columns: list[str] | None = None,
    ) -> str:
        """
        Створює файл із заголовками з налаштувань / аргументу або дописує рядок.
        Шлях файлу — ``self.file_path``. Якщо ``export_columns`` не передано — з ``ConfigManager``.
        """
        cfg_cols = export_columns
        if cfg_cols is None:
            cfg_cols = list(self._config_mgr.load().get("export_columns") or DEFAULT_SETTINGS["export_columns"])

        cols, row = self.build_row_for_columns(data, cfg_cols)
        path = self.file_path

        def _save_workbook(workbook: Workbook) -> None:
            try:
                workbook.save(path)
            except (PermissionError, OSError) as exc:
                name = os.path.basename(path)
                raise PermissionError(
                    f"Не вдалося записати «{name}». Можливі причини: файл відкритий у Excel або іншій "
                    f"програмі; папку блокує OneDrive / антивірус; немає прав на запис. Деталі: {exc!r}"
                ) from exc

        if not os.path.isfile(path):
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(cols)
            ws.append(row)
            _save_workbook(wb)
        else:
            try:
                wb = load_workbook(path)
            except (PermissionError, OSError) as exc:
                name = os.path.basename(path)
                raise PermissionError(
                    f"Не вдалося відкрити «{name}» для допису. Перевірте, чи файл не відкритий у Excel, "
                    f"чи не блокує доступ OneDrive/антивірус. Деталі: {exc!r}"
                ) from exc
            ws = wb.active
            assert ws is not None
            ws.append(row)
            _save_workbook(wb)

        return os.path.abspath(path)
